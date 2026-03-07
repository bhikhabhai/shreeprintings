import openpyxl

wb = openpyxl.load_workbook('d:/Rudra/AttendanceSheet_Full.xlsx', data_only=True)
ws = wb[wb.sheetnames[0]]

# Data rows 3-5 for cols 6-10 (first date: 01-Jan)
print("=== DATA VALUES for 01-Jan (cols 6-10) ===")
for r in range(3, min(ws.max_row + 1, 8)):
    emp_id = ws.cell(r, 1).value
    print(f"Row {r} ({emp_id}):")
    for c in range(6, 11):
        v = ws.cell(r, c).value
        t = type(v).__name__ if v is not None else "None"
        header = ws.cell(2, c).value
        print(f"  {header} (col {c}): {repr(v)} [{t}]")

# All employee rows info
print("\n=== ALL EMPLOYEES ===")
for r in range(3, ws.max_row + 1):
    vals = [ws.cell(r, c).value for c in range(1, 6)]
    print(f"  Row {r}: {vals}")

# Row 1 date headers - collect them
print("\n=== ALL DATE HEADERS (Row 1) ===")
for c in range(6, 161):
    v = ws.cell(1, c).value
    if v is not None:
        print(f"  Col {c}: {repr(v)}")
